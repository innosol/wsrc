# -*- coding: utf-8 -*-
import jsonpickle
import openpyxl
import re
import xlsxwriter
import StringIO
from django.contrib.messages.views import SuccessMessageMixin
from django.core.urlresolvers import reverse_lazy
from django.shortcuts import render_to_response
from django.views.generic import FormView, TemplateView
from django.http import HttpResponse, Http404, HttpResponseRedirect
from applications.app.views import LoginRequired
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .forms import (SezForm, OlborloltForm , ZardalForm16, Tehnik_nohtsolForm, BusadOrlogoForm)
from django.utils import timezone
from .models import *
from applications.app.models import User, TZE
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def converter(value):
	#a = re.sub('[,]', '', value)
	return float(value)

class TailanView(object):
	tailan = None
	date = None
	user = None
	baiguullaga = None
	string = ''
	model = None
	cls = None

	def dispatch(self, request, *args, **kwargs):
		try:
			self.user = User.objects.get(id = request.user.id)
			self.baiguullaga = TZE.objects.get(id = self.user.user_id.baiguullaga.id)
			if 'date' in self.kwargs:
				self.tailan = SariinTailan.objects.get(id = self.kwargs['date'])
				#self.tailan = Sudalgaa.objects.get(tze  = self.baiguullaga, year = self.kwargs['date'])
				self.date = self.kwargs['date']
			else:
				self.tailan = SariinTailan.objects.get(tze  = self.baiguullaga, tailan_status = True)
		except: pass
		try:
			if hasattr(self.tailan.zardal, self.string):
				self.model = getattr(self.tailan.zardal, self.string)
				self.cls = getattr(self.model, self.model._meta.get_m2m_with_model()[0][0].name)
		except: pass
		return super(TailanView, self).dispatch(request, *args, **kwargs)

	def get_context_data(self, *args, **kwargs):
		context = super(TailanView, self).get_context_data(*args, **kwargs)
		context['tailan'] = self.tailan
		context['date'] = self.date
		context['hereglegch'] = HereglegchAll.objects.filter(tze = self.baiguullaga).last()
		context['golch'] =Golch.objects.filter(tze = self.baiguullaga, name = 0).last()
		context['tariffgolch'] = TariffAll.objects.filter(tze = self.baiguullaga, name = 1).last()
		context['golch_hangagch'] = Golch.objects.filter(tze = self.baiguullaga, name = 1).last()
		context['tariffus'] = TariffAll.objects.filter(tze = self.baiguullaga, name =0).last()
		context['ol'] = Olborlolt.objects.filter(tze = self.baiguullaga).last()
		context['sudalgaa'] =  SariinTailan.objects.filter(tze = self.baiguullaga, year = 2015, yvts__in = [u'Хийгдэж байна', u'Буцаасан']).order_by('month') #Sudalgaa.objects.filter(tze = self.baiguullaga)
		if context['tariffus'] and context['tariffgolch'] and context['hereglegch'] and context['ol'] and context['golch'] and context['golch_hangagch']:
			context['should'] = True
		try:
			if SariinTailan.objects.filter(tze = self.tailan.tze, year = self.tailan.year, month = self.tailan.month-1, yvts = u'Хүлээн авсан') and self.tailan.yvts != u'Илгээсэн':
				context['umnuh_tailan'] = SariinTailan.objects.get(tze = self.tailan.tze, year = self.tailan.year, month = self.tailan.month-1, yvts = u'Хүлээн авсан').id
		except: pass
		return context

class EasyFormClass(TailanView):

	def form_valid(self, form):
		if self.model:
			for too, i in enumerate(self.cls.all()):
				count = 0
				for l in self.model._meta.get_m2m_with_model()[0][0].rel.to.catch():
					setattr(i, l, form.cleaned_data[self.attr %str(too+count)])
					count += 7
				i.save()
		else:	
			z_many = self.many_object.objects.create(
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True,
				)
			z_simple = getattr(z_many, z_many._meta.get_m2m_with_model()[0][0].name)
			for too in range(7):
				z = self.simple_object(
					torol = too
					)
				count = 0
				for l in z_many._meta.get_m2m_with_model()[0][0].rel.to.catch():
					setattr(z, l, form.cleaned_data[self.attr %str(too+count)])
					count += 7
				z.save()
				z_simple.add(z)
			z_many.save()
			setattr(self.tailan.zardal, self.string, z_many)
			self.tailan.zardal.save()
			self.model = getattr(self.tailan.zardal, self.string, z_many)
		self.model.totals()
		return super(TailanView, self).form_valid(form)

	def get_initial(self):
		initial = super(TailanView, self).get_initial()
		try:
			for too, i in enumerate(self.cls.all()):
				count = 0
				for l in self.model._meta.get_m2m_with_model()[0][0].rel.to.catch():
					initial[self.attr %str(too+count)] = getattr(i, l)
					count += 7
		except: pass
		return initial

class SezView(LoginRequired, TailanView, TemplateView):
	template_name = 'sez.html'
	
	def get_context_data(self, **kwargs):
		context = super(SezView, self).get_context_data(**kwargs)
		ilgeesen = butsaasan = hiigdehgui = 0
		hiigdej_bui = 25
		if self.tailan:
			if hasattr(self.tailan, 'tasaldal'):
				if self.tailan.tasaldal != None:
					ilgeesen += 1
			if hasattr(self.tailan, 'tehnik_nohtsol'):
				if self.tailan.tehnik_nohtsol != None:
					ilgeesen += 1
			if hasattr(self.tailan, 'sanal_gomdol'):
				if self.tailan.sanal_gomdol != None:
					ilgeesen += 1
			if hasattr(self.tailan, 'tsalin'):
				if self.tailan.tsalin != None:
					ilgeesen += 1
		if context['ol']:
			if context['ol'].torol1 == True and context['ol'].torol2 == False:
				hiigdehgui+=1
		if hasattr(self.tailan, 'zardal'):
			for i in range(1, 20):
				if hasattr(self.tailan.zardal, 'z%s' %i):
					if getattr(self.tailan.zardal, 'z%s' %i) != None:
						if getattr(self.tailan.zardal, 'z%s' %i).yvts == u'Хадгалсан':
							ilgeesen+=1
						if getattr(self.tailan.zardal, 'z%s' %i).yvts == u'Буцаасан':
							butsaasan+=1
		if hasattr(self.tailan, 'orlogo'):
			if self.tailan.orlogo.buteegdehuun.all():
				ilgeesen +=1

		if hasattr(self.tailan, 'busad_orlogo'):
			if getattr(self.tailan, 'busad_orlogo') != None:
				if getattr(self.tailan, 'busad_orlogo').yvts == u'Хадгалсан':
					ilgeesen += 1

		context['ilgeesen'] = ilgeesen
		context['butsaasan'] = butsaasan
		context['hiigdehgui'] = hiigdehgui
		context['hiigdej_bui'] = int(hiigdej_bui)-int(ilgeesen)-int(butsaasan)-int(hiigdehgui)
		paginator = Paginator(SariinTailan.objects.filter(tze = self.baiguullaga, yvts__in = [u'Илгээсэн', u'Буцаасан', u'Хүлээн авсан'], month__gt = 0, status = True), 10)
		page = self.request.GET.get('page')
		try: context['sariin_tailan'] = paginator.page(page)
		except PageNotAnInteger: context['sariin_tailan'] = paginator.page(1)
		except EmptyPage: context['sariin_tailan'] = paginator.page(paginator.num_pages)
		return context

	@staticmethod
	def copy(request, id = 0, tid = 0):
		a = SariinTailan.objects.get(id = id)
		a.orlogo.buteegdehuun.clear()
		b = SariinTailan.objects.get(id = tid)
		
		# busad orlogo tatah
		busad_orlogo = b.busad_orlogo
		busad_orlogo.id = None
		busad_orlogo.save()
		a.busad_orlogo = busad_orlogo
		# usnii bichilt tatah
		a.orlogo.olborloson_us = b.orlogo.olborloson_us
		a.orlogo.save()
		for i in b.orlogo.buteegdehuun.all():
			i.id = None
			i.save()
			a.orlogo.buteegdehuun.add(i)
		a.orlogo.yvts = u'Хадгалсан'
		a.orlogo.save()
		# tsalin tatah
		tsalin = TsalinMany.objects.create(yvts = u'Хадгалсан')
		for s in b.tsalin.tsalin.all():
			s.id = None
			s.save()
			tsalin.tsalin.add(s)
		a.tsalin = tsalin
		# Sanal gomdol tatah
		sanal_gomdol = SanalGomdolMany.objects.create(tze = a.tze, yvts = u'Хадгалсан')
		for i in b.sanal_gomdol.sanal_gomdol.all():
			i.id = None
			i.save()
			sanal_gomdol.sanal_gomdol.add(i)
		sanal_gomdol.totals()
		a.sanal_gomdol = sanal_gomdol
		# Tehnik nohtsol tatah
		tehnik_nohtsol = TehnikNohtsolMany.objects.create(tze = b.tze, yvts = u'Хадгалсан')
		for i in b.tehnik_nohtsol.tehnik_nohtsol.all():
			i.id = None
			i.save()
			tehnik_nohtsol.tehnik_nohtsol.add(i)
		tehnik_nohtsol.totals()
		a.tehnik_nohtsol = tehnik_nohtsol
		# tasaldal tatah
		tasaldal = TasaldalMany.objects.create(tze = b.tze, yvts = u'Хадгалсан')
		for i in b.tasaldal.tasaldal.all():
			i.id = None
			i.save()
			tasaldal.tasaldal.add(i)
		tasaldal.totals()
		a.tasaldal = tasaldal
		a.save()
		# zardal tatah
		for i in range(1,20):
			z = getattr(b.zardal, "z%s" %i)
			if z == None:
				continue
			else:
				zz = z.__class__.objects.create(yvts = u'Хадгалсан')
				many_field = getattr(z, z._meta.get_m2m_with_model()[0][0].name)
				many_field1 = getattr(zz, zz._meta.get_m2m_with_model()[0][0].name)
				if int(i) == 16:
					for s in many_field.all():
						busad_middle = ZardalBusadMiddle.objects.create(name = s.name)
						busad_middle.counts()
						for ss in s.busad.all():
							ss.id = None
							ss.save()
							busad_middle.busad.add(ss)
						many_field1.add(busad_middle)
					zz.save()
					setattr(a.zardal, 'z%s' %i, zz)

				else:
					for s in many_field.all():
						s.id = None
						s.save()
						many_field1.add(s)
					zz.totals()
					setattr(a.zardal, 'z%s' %i, zz)
		a.zardal.save()
		return HttpResponseRedirect('/director/sez/%s/' %id)

class ZardalView(LoginRequired, SuccessMessageMixin, FormView):
	success_message = u"Таны мэдээлэл амжилттай хадгалагдлаа"
	success_url = reverse_lazy('sez')

	def get_success_url(self, **kwargs):
		if self.date:
			return reverse_lazy('sudalgaa', kwargs = {'date' : self.kwargs['date']})
		else:
			return reverse_lazy('sez')

class SezZardalView(ZardalView):
	number = 49
	attr = 'z2_%s'
	integer = False
	form_class = SezForm

	def get_form_kwargs(self):
		kwargs = super(SezZardalView, self).get_form_kwargs()
		kwargs['number'] = self.number
		kwargs['attr'] = self.attr
		kwargs['integer'] = self.integer
		return kwargs

class OrlogoFormView(TailanView, SezZardalView):
	template_name = 'formuud/orlogo.html'
	number = 19
	attr = 'fields_%s'

	def form_valid(self, form):
		if self.tailan.orlogo.buteegdehuun.all():
			for too, i in enumerate(self.tailan.orlogo.buteegdehuun.all()):
				if too < 19:
					i.too = converter(form.cleaned_data['fields_%s' %too])
					i.save()
				else:
					break
			self.tailan.orlogo.olborloson_us = form.cleaned_data['fields_18']
			self.tailan.orlogo.save()
		else:
			too=0
			for i in range(9):
				b = Buteegdehuun.objects.create(
					too = converter(form.cleaned_data['fields_%s' %too]), angilal = 0, name = i)
				b1 = Buteegdehuun.objects.create(
					too = converter(form.cleaned_data['fields_%s' %str(too+1)]), angilal = 1, name = i)
				too+=2
				self.tailan.orlogo.buteegdehuun.add(b, b1)
				self.tailan.orlogo.yvts = u'Хадгалсан'
			self.tailan.orlogo.olborloson_us = form.cleaned_data['fields_18']
			self.tailan.orlogo.save()
		return super(OrlogoFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(OrlogoFormView, self).get_initial()
		try:
			for too, i in enumerate(self.tailan.orlogo.buteegdehuun.all()):
				initial['fields_%s' %(too)] = i.too
			initial['fields_18'] = self.tailan.orlogo.olborloson_us
		except: pass
		return initial

class BusadOrlogoFormView(TailanView, ZardalView):
	form_class = BusadOrlogoForm
	template_name = 'formuud/busadorlogo.html'

	def form_valid(self, form):
		if self.tailan.busad_orlogo:
			setattr(self.tailan.busad_orlogo, 'unt', form.cleaned_data['unt'])
			for i in range(6):
				setattr(self.tailan.busad_orlogo, 'unt%s' %i, form.cleaned_data['unt%s' %i])
			self.tailan.busad_orlogo.save()

		else:
			objects = form.save(commit = False)
			objects.status = True
			objects.created_by = self.user
			objects.yvts = u'Хадгалсан'
			objects.save()
			self.tailan.busad_orlogo = objects
			self.tailan.save()
		return super(BusadOrlogoFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(BusadOrlogoFormView, self).get_initial()
		try:
			initial['unt'] = getattr(self.tailan.busad_orlogo, 'unt')
			for i in range(6):
				initial['unt%s' %i] = getattr(self.tailan.busad_orlogo, 'unt%s' %i)
		except: pass
		return initial

class TariffFormView(TailanView, SezZardalView):
	template_name = 'formuud/tariff.html'
	number = 17
	attr = 'su_%s'

	def form_valid(self, form):
		ta = TariffAll.objects.create(
				tze = self.baiguullaga,
				name = 0,
				created_by = self.user,
				status = True,
				suuri_une = form.cleaned_data['su_16']
				)
		too=0
		for i in range(8):
			t = Tariff.objects.create(une = converter(form.cleaned_data['su_%s' %too]), angilal = 0, name = i+1)
			t1 = Tariff.objects.create(une = converter(form.cleaned_data['su_%s' %str(too+1)]), angilal = 1, name = i+1)
			too+=2
			ta.tariff.add(t, t1)
		return super(TariffFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(TariffFormView, self).get_initial()
		try:
			tariff = TariffAll.objects.filter(tze = self.baiguullaga, name =0).last()
			for too in range(16):
				initial['su_%s' %(too)] =tariff.tariff.all()[too].une
			initial['su_16'] = tariff.suuri_une
		except:
			pass
		return initial

class HereglegchFormview(TailanView, SezZardalView):
	template_name = 'formuud/hereglegch.html'
	number = 18
	attr = 'su_%s'
	integer = True
	
	def form_valid(self, form):
		ta = HereglegchAll.objects.create(
				tze = self.baiguullaga,
				created_by = self.user,
				status = True
				)
		too=0
		for i in range(9):
			t = Hereglegch.objects.create(htoo = form.cleaned_data['su_%s' %too], angilal = 0, name = i+1)
			t1 = Hereglegch.objects.create(htoo = form.cleaned_data['su_%s' %str(too+1)], angilal = 1, name = i+1)
			too+=2
			ta.hereglegch.add(t, t1)
		return super(HereglegchFormview, self).form_valid(form)

	def get_initial(self):
		initial = super(HereglegchFormview, self).get_initial()
		try:
			hereglegch = HereglegchAll.objects.filter(tze = self.baiguullaga).last().hereglegch
			for i in range(18):
				initial['su_%s' %(i)] = hereglegch.all()[i].htoo
		except: pass
		return initial

class UsHudaldanAvahFormView(TailanView, ZardalView):
	
	template_name = "formuud/ushudaldanavah.html"
	form_class = OlborloltForm

	def form_valid(self, form):
		o = form.save(commit = False)
		o.tze = self.baiguullaga
		o.created_by = self.user
		o.save()
		return super(UsHudaldanAvahFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(UsHudaldanAvahFormView, self).get_initial()
		try:
			initial['torol1'] = Olborlolt.objects.filter(tze = self.baiguullaga).last().torol1
			initial['torol2'] = Olborlolt.objects.filter(tze = self.baiguullaga).last().torol2
			initial['tsever'] = Olborlolt.objects.filter(tze = self.baiguullaga).last().tsever
			initial['bohir'] = Olborlolt.objects.filter(tze = self.baiguullaga).last().bohir
		except:
			pass
		return initial

class AANBFormView(TailanView, SezZardalView):
	template_name = 'formuud/golchiinsuuriune.html'
	number = 15
	attr = 'golch_%s'

	def form_valid(self, form):
		ta = TariffAll.objects.create(tze = self.baiguullaga, status = True, name = 1, created_by = self.user)
		for i in range(15):
			t = Tariff.objects.create(angilal = 0, name = i+9, une = converter(form.cleaned_data['golch_%s' %i]))
			ta.tariff.add(t)
		return super(AANBFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(AANBFormView, self).get_initial()
		try:
			for i in range(15):
				initial['golch_%s' %i] = TariffAll.objects.filter(tze = self.baiguullaga, name = 1).last().tariff.all()[i].une
		except: pass
		return initial

class GolchFormView(TailanView, SezZardalView):
	template_name = 'formuud/golch.html'
	number = 15
	attr = 'golch_%s'
	integer = True

	def form_valid(self, form):
		ta = Golch.objects.create(tze = self.baiguullaga, name = 0, created_by = self.user, status = True)
		for i in range(15):
			t = GolchToo.objects.create(
				songolt = i+9,
				too = form.cleaned_data['golch_%s' %i],
				)
			ta.buteegdehuun.add(t)
		ta.totals()
		return super(GolchFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(GolchFormView, self).get_initial()
		try:
			for i in range(15):
				initial['golch_%s' %i] = Golch.objects.filter(tze = self.baiguullaga, name = 0).last().buteegdehuun.all()[i].too
		except:
			pass
		return initial

class GolchHangagchFormView(TailanView, SezZardalView):
	
	template_name = 'formuud/golch.html'
	number = 15
	attr = 'golch_%s'
	integer = True

	def form_valid(self, form):
		ta = Golch.objects.create(tze = self.baiguullaga, name = 1, created_by = self.user, status = True)
		for i in range(15):
			t = GolchToo.objects.create(
				songolt = i+9,
				too = form.cleaned_data['golch_%s' %i]
				)
			ta.buteegdehuun.add(t)
		ta.totals()
		return super(GolchHangagchFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(GolchHangagchFormView, self).get_initial()
		try:
			for i in range(15):
				initial['golch_%s' %i] = Golch.objects.filter(tze = self.baiguullaga, name = 1).last().buteegdehuun.all()[i].too
		except:
			pass
		return initial

class ZardalFormView1(TailanView, SezZardalView):
	number = 7
	attr = 'z1_%s'
	template_name = 'formuud/zardal1.html'
	
	def form_valid(self, form):
		if self.tailan.zardal.z1:
			for too, i in enumerate(self.tailan.zardal.z1.undsen_material.all()):
				i.undsen_tuuhii_ed_usnii = form.cleaned_data[self.attr %too]
				i.save()
		else:
			z_many = ZardalUndsenMaterialMany.objects.create(yvts = u'Хадгалсан', created_by = self.user)
			for i in range(7):
				z = ZardalUndsenMaterial.objects.create(undsen_tuuhii_ed_usnii = form.cleaned_data[self.attr % i], torol = i)
				z_many.undsen_material.add(z)
			z_many.save()
			self.tailan.zardal.z1 = z_many
			self.tailan.zardal.save()
		return super(ZardalFormView1, self).form_valid(form)

	def get_initial(self):
		initial = super(ZardalFormView1, self).get_initial()
		try:
			for too, i in enumerate(self.tailan.zardal.z1.undsen_material.all()):
				initial[self.attr %too] = i.undsen_tuuhii_ed_usnii
		except: pass
		return initial

class ZardalFormView2(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal2.html'
	attr = 'z2_%s'
	number = 21
	string = 'z2'
	many_object = ZardalTsalinMany
	simple_object = ZardalTsalin

class ZardalFormView3(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal3.html'
	number = 49
	attr = 'z3_%s'
	string = 'z3'
	many_object = ZardalAshiglaltMany
	simple_object = ZardalAshiglalt

class ZardalFormView4(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal4.html'
	number = 42
	attr = 'z4_%s'
	string = 'z4'
	many_object = ZardalZasvarUilchilgeeMany
	simple_object = ZardalZasvarUilchilgee

class ZardalFormView5(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal5.html'
	number = 28
	attr = 'z5_%s'
	string = 'z5'
	many_object = ZardalAriutgalMany
	simple_object = ZardalAriutgal

class ZardalFormView6(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal6.html'
	number = 35
	attr = 'z6_%s'
	string = 'z6'
	many_object = ZardalKontorMany
	simple_object = ZardalKontor

class ZardalFormView7(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal7.html'
	number = 28
	attr = 'z7_%s'
	string = 'z7'
	many_object = ZardalHodolmorHamgaalalMany
	simple_object = ZardalHodolmorHamgaalal

class ZardalFormView8(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal8.html'
	number = 35
	attr = 'z8_%s'
	string = 'z8'
	many_object = ZardalMarketingMany
	simple_object = ZardalMarketing

class ZardalFormView9(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal9.html'
	number = 35
	attr = 'z9_%s'
	string = 'z9'
	many_object = ZardalLaboratoryMany
	simple_object = ZardalLaboratory

class ZardalFormView10(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal10.html'
	number = 35
	attr = 'z10_%s'
	string = 'z10'
	many_object = ZardalGuitsetgehUdirdlagaMany
	simple_object = ZardalGuitsetgehUdirdlaga

class ZardalFormView11(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal11.html'
	number = 21
	attr = 'z11_%s'
	string = 'z11'
	many_object = ZardalTUZMany
	simple_object = ZardalTUZ

class ZardalFormView12(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal12.html'
	number = 28
	attr = 'z12_%s'
	string = 'z12'
	many_object = ZardalUndsenHorongiinElegdelMany
	simple_object = ZardalUndsenHorongiinElegdel

class ZardalFormView13(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal13.html'
	number = 49
	attr = 'z13_%s'
	string = 'z13'
	many_object = ZardalGadniiUilchilgeeMany
	simple_object = ZardalGadniiUilchilgee

class ZardalFormView14(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal14.html'
	number = 28
	attr = 'z14_%s'
	string = 'z14'
	many_object = ZardalTatvarMany
	simple_object = ZardalTatvar

class ZardalFormView15(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal15.html'
	number = 28
	attr = 'z15_%s'
	string = 'z15'
	many_object = ZardalDaatgalMany
	simple_object = ZardalDaatgal

class ZardalFormView16(TailanView, ZardalView):
	template_name = 'formuud/zardal16.html'
	form_class = ZardalForm16

	def form_valid(self, form):
		name = self.request.POST.getlist('z16_0')
		v1 = self.request.POST.getlist('z16_1')
		v2 = self.request.POST.getlist('z16_2')
		v3 = self.request.POST.getlist('z16_3')
		v4 = self.request.POST.getlist('z16_4')
		v5 = self.request.POST.getlist('z16_5')
		v6 = self.request.POST.getlist('z16_6')
		v7 = self.request.POST.getlist('z16_7')
		
		if len(name) == len(v1) and len(v2) == len(v3) and len(v4) == len(v5) and\
		len(name) == len(v6) and len(v6) == len(v2) and len(v3) == len(v4):
			if self.tailan.zardal.z16:
				z_many = self.tailan.zardal.z16
				for zardal in z_many.busad_zardal.all():
					zardal.busad.all().delete()
					zardal.save() 
				z_many.busad_zardal.all().delete()
				z_many.save()
				for i in range(len(name)):
					z1 = ZardalBusadZardal.objects.create(value = v1[i], torol = 0)
					z2 = ZardalBusadZardal.objects.create(value = v2[i], torol = 1)
					z3 = ZardalBusadZardal.objects.create(value = v3[i], torol = 2)
					z4 = ZardalBusadZardal.objects.create(value = v4[i], torol = 3)
					z5 = ZardalBusadZardal.objects.create(value = v5[i], torol = 4)
					z6 = ZardalBusadZardal.objects.create(value = v6[i], torol = 5)
					z7 = ZardalBusadZardal.objects.create(value = v7[i], torol = 6)
					z_middle = ZardalBusadMiddle.objects.create(name = name[i])
					z_middle.busad.add(z1)
					z_middle.busad.add(z2)
					z_middle.busad.add(z3)
					z_middle.busad.add(z4)
					z_middle.busad.add(z5)
					z_middle.busad.add(z6)
					z_middle.busad.add(z7)
					z_middle.counts()
					z_many.busad_zardal.add(z_middle)
					z_many.save()
			else:
				z_many = ZardalBusadZardalMany.objects.create(
					created_by = self.user,
					yvts = u'Хадгалсан',
					status = True,
					)
				for i in range(len(name)):
					z1 = ZardalBusadZardal.objects.create(value = v1[i], torol = 0)
					z2 = ZardalBusadZardal.objects.create(value = v2[i], torol = 1)
					z3 = ZardalBusadZardal.objects.create(value = v3[i], torol = 2)
					z4 = ZardalBusadZardal.objects.create(value = v4[i], torol = 3)
					z5 = ZardalBusadZardal.objects.create(value = v5[i], torol = 4)
					z6 = ZardalBusadZardal.objects.create(value = v6[i], torol = 5)
					z7 = ZardalBusadZardal.objects.create(value = v7[i], torol = 6)
					z_middle = ZardalBusadMiddle.objects.create(
						name = name[i],
						)
					z_middle.busad.add(z1)
					z_middle.busad.add(z2)
					z_middle.busad.add(z3)
					z_middle.busad.add(z4)
					z_middle.busad.add(z5)
					z_middle.busad.add(z6)
					z_middle.busad.add(z7)
					z_middle.counts()
					z_many.busad_zardal.add(z_middle)
					z_many.save()
			z_many.counts()
			self.tailan.zardal.z16 = z_many
			self.tailan.zardal.save()
		return super(ZardalFormView16, self).form_valid(form)

class ZardalFormView17(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal17.html'
	number = 35
	attr = 'z17_%s'
	string = 'z17'
	many_object = ZardalAjilchidNiigmiinMany
	simple_object = ZardalAjilchidNiigmiin

class ZardalFormView18(EasyFormClass, SezZardalView):
	template_name = 'formuud/zardal18.html'
	number = 14
	attr = 'z18_%s'
	string = 'z18'
	many_object = ZardalHorongoOruulaltMany
	simple_object = ZardalHorongoOruulalt

class ZardalFormView19(TailanView, SezZardalView):
	template_name = 'formuud/zardal19.html'
	number = 12
	attr = 'z19_%s'

	def form_valid(self, form):
		if self.tailan.zardal.z19:
			for too, i in enumerate(self.tailan.zardal.z19.undsen_bus_uil_ajillagaa.all()):
				i.shagnal_uramshuulal = converter(form.cleaned_data['z19_%s' %str(too)])
				i.sport_soyliin_arga_hemjee = converter(form.cleaned_data['z19_%s' %str(too+4)])
				i.busad = converter(form.cleaned_data['z19_%s' %str(too+8)])
				i.save()
			pass	
		else:
			z_many = ZardalUndsenBusUilAjillagaaMany.objects.create(
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True
				)
			for too in range(4):
				z = ZardalUndsenBusUilAjillagaa.objects.create(
					torol = too,
					shagnal_uramshuulal = converter(form.cleaned_data['z19_%s' %str(too)]),
					sport_soyliin_arga_hemjee = converter(form.cleaned_data['z19_%s' %str(too+4)]),
					busad = converter(form.cleaned_data['z19_%s' %str(too+8)])
					)
				z_many.undsen_bus_uil_ajillagaa.add(z)
			z_many.save()
			self.tailan.zardal.z19 = z_many
			self.tailan.zardal.save()
		self.tailan.zardal.z19.totals()
		return super(ZardalFormView19, self).form_valid(form)

	def get_initial(self):
		initial = super(ZardalFormView19, self).get_initial()
		try:
			for i in range(4):
				initial['z19_%s' %i] = self.tailan.zardal.z19.undsen_bus_uil_ajillagaa.all()[i].shagnal_uramshuulal
				initial['z19_%s' %str(i+4)] = self.tailan.zardal.z19.undsen_bus_uil_ajillagaa.all()[i].sport_soyliin_arga_hemjee
				initial['z19_%s' %str(i+8)] = self.tailan.zardal.z19.undsen_bus_uil_ajillagaa.all()[i].busad
		except:
			pass
		return initial		

class TsalinFormView(TailanView, SezZardalView):

	template_name = 'formuud/tsalin.html'
	number = 15
	attr = 'tsalin_%s'

	def form_valid(self, form):
		if self.tailan.tsalin:
			num = 0
			for i in self.tailan.tsalin.tsalin.all():
				i.too = form.cleaned_data['tsalin_%s' %num]
				i.undsen_tsalin = form.cleaned_data['tsalin_%s' %str(num+1)]
				i.nemegdel = form.cleaned_data['tsalin_%s' %str(num+2)]
				i.sariin_ur_dun = form.cleaned_data['tsalin_%s' %str(num+3)]
				i.jiliin_ur_dun_shagnalt_tsalin = form.cleaned_data['tsalin_%s' %str(num+4)]
				i.save()
				num += 5
		else:
			tsalin = TsalinMany.objects.create(
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True
				)
			for t, i in enumerate(range(0, 15, 5)):
				z = Tsalin.objects.create(
					songolt = i,
					too = form.cleaned_data['tsalin_%s' %i],
					undsen_tsalin = form.cleaned_data['tsalin_%s' %str(i+1)],
					nemegdel = form.cleaned_data['tsalin_%s' %str(i+2)],
					sariin_ur_dun = form.cleaned_data['tsalin_%s' %str(i+3)],
					jiliin_ur_dun_shagnalt_tsalin = form.cleaned_data['tsalin_%s' %str(i+4)]
					)
				tsalin.tsalin.add(z)
			self.tailan.tsalin = tsalin
			self.tailan.save()
		return super(TsalinFormView, self).form_valid(form)

	def get_initial(self):
		initial = super(TsalinFormView, self).get_initial()
		try:
			num = 0
			for z in self.tailan.tsalin.tsalin.all():
				initial['tsalin_%s' %num] = z.too
				initial['tsalin_%s' %str(num+1)] = z.undsen_tsalin
				initial['tsalin_%s' %str(num+2)] = z.nemegdel
				initial['tsalin_%s' %str(num+3)] = z.sariin_ur_dun
				initial['tsalin_%s' %str(num+4)] = z.jiliin_ur_dun_shagnalt_tsalin
				num += 5
		except: pass
		return initial

class SanalGomdolView(TailanView, SezZardalView):
	template_name = 'formuud/sanalgomdol.html'
	number = 10
	attr = 'sg_%s'
	integer = True

	def form_valid(self, form):
		num = 0
		if self.tailan.sanal_gomdol:
			for z in self.tailan.sanal_gomdol.sanal_gomdol.all():
				z.huleen_avsan = form.cleaned_data['sg_%s' %str(num)]
				z.shiidverlesen = form.cleaned_data['sg_%s' %str(num+1)]
				z.save()
				num +=2
		else:
			z_many = SanalGomdolMany.objects.create(
				tze = self.baiguullaga,
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True,
				)
			for too in range(5):
				z = SanalGomdol.objects.create(
					torol = too,
					huleen_avsan = form.cleaned_data['sg_%s' %str(num)],
					shiidverlesen = form.cleaned_data['sg_%s' %str(num+1)]
					)
				num +=2
				z_many.sanal_gomdol.add(z)
			self.tailan.sanal_gomdol = z_many
			self.tailan.save()
		self.tailan.sanal_gomdol.totals()
		return super(SanalGomdolView, self).form_valid(form)

	def get_initial(self):
		initial = super(SanalGomdolView, self).get_initial()
		try:
			num = 0
			for z in self.tailan.sanal_gomdol.sanal_gomdol.all():
				initial['sg_%s' %num] = z.huleen_avsan
				initial['sg_%s' %str(num+1)] = z.shiidverlesen
				num += 2
		except: pass
		return initial

class TehnikNohtsolView(TailanView, ZardalView):
	template_name = 'formuud/tehnik_nohtsol.html'
	form_class = Tehnik_nohtsolForm

	def form_valid(self, form):
		num = 0
		if self.tailan.tehnik_nohtsol:
			for i in self.tailan.tehnik_nohtsol.tehnik_nohtsol.all():
				i.too = form.cleaned_data['tn_%s' %(num)]
				i.us = form.cleaned_data['tn_%s' %(num+1)]
				i.save()
				num += 2
		else:
			z_many = TehnikNohtsolMany.objects.create(
				tze = self.baiguullaga,
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True
				)
			for i in range(2):
				z = TehnikNohtsol.objects.create(
					torol = i,
					too = form.cleaned_data['tn_%s' %(num)],
					us = form.cleaned_data['tn_%s' %(num+1)]
					)
				num += 2
				z_many.tehnik_nohtsol.add(z)
			self.tailan.tehnik_nohtsol = z_many
			self.tailan.save()
		self.tailan.tehnik_nohtsol.totals()
		return super(TehnikNohtsolView, self).form_valid(form)

	def get_initial(self):
		initial = super(TehnikNohtsolView, self).get_initial()
		try:
			initial['tn_0'] = self.tailan.tehnik_nohtsol.tehnik_nohtsol.all()[0].too
			initial['tn_1'] = self.tailan.tehnik_nohtsol.tehnik_nohtsol.all()[0].us
			initial['tn_2'] = self.tailan.tehnik_nohtsol.tehnik_nohtsol.all()[1].too
			initial['tn_3'] = self.tailan.tehnik_nohtsol.tehnik_nohtsol.all()[1].us
		except:
			pass
		return initial

class TasaldalView(TailanView, SezZardalView):
	template_name = 'formuud/tasaldal.html'
	number = 14
	attr = 't_%s'
	integer = True

	def form_valid(self, form):
		num = 0
		if self.tailan.tasaldal:
			for i in self.tailan.tasaldal.tasaldal.all():
				i.duration = form.cleaned_data['t_%s' %(num)]
				i.too = form.cleaned_data['t_%s' %(num+1)]
				i.save()
				num += 2
		else:
			z_many = TasaldalMany.objects.create(
				tze = self.baiguullaga,
				yvts = u'Хадгалсан',
				created_by = self.user,
				status = True
				)
			for i in range(7):
				z = Tasaldal.objects.create(
					torol = i,
					duration = form.cleaned_data['t_%s' %(num)],
					too = form.cleaned_data['t_%s' %(num+1)],
					)
				num += 2
				z_many.tasaldal.add(z)
			self.tailan.tasaldal = z_many
			self.tailan.save()
		self.tailan.tasaldal.totals()
		return super(TasaldalView, self).form_valid(form)

	def get_initial(self):
		initial = super(TasaldalView, self).get_initial()
		num = 0
		try:
			for i in self.tailan.tasaldal.tasaldal.all():
				initial['t_%s' %(num)] = i.duration
				initial['t_%s' %(num+1)] = i.too
				num +=2
		except:
			pass
		return initial











def sendtailan(request, id = 0):
	a = SariinTailan.objects.get(id =id)
	tze = a.tze
	
	a.olborlolt = Olborlolt.objects.filter(tze = tze).last()
	a.hereglegch = HereglegchAll.objects.filter(tze = tze).last()
	a.tariff_hereglegch.clear()
	a.golch.clear()
	a.tariff_hereglegch.add(TariffAll.objects.filter(tze = tze, name =0).last())
	a.tariff_hereglegch.add(TariffAll.objects.filter(tze = tze, name =1).last())
	a.golch.add(Golch.objects.filter(tze = tze, name = 0).last())
	a.golch.add(Golch.objects.filter(tze = tze, name = 1).last())	
	a.tailan_status = False
	a.yvts = u'Илгээсэн'
	a.save()
	return HttpResponseRedirect('/director/sez/')

def sendsudalgaa(request, id = 0):
	a = Sudalgaa.objects.get(id =id)
	tze = a.tze
	a.olborlolt = Olborlolt.objects.filter(tze = tze).last()
	a.tariff_hereglegch.clear()
	a.golch.clear()
	a.hereglegch = HereglegchAll.objects.filter(tze = tze).last()
	a.tariff_hereglegch.add(TariffAll.objects.filter(tze = tze, name =0).last())
	a.tariff_hereglegch.add(TariffAll.objects.filter(tze = tze, name =1).last())
	a.golch.add(Golch.objects.filter(tze = tze, name = 0).last())
	a.golch.add(Golch.objects.filter(tze = tze, name = 1).last())	
	a.yvts = u'Илгээсэн'
	a.save()
	return HttpResponseRedirect('/director/sez/')
'''
def Export(request):
	xfile = openpyxl.load_workbook('test.xlsx')
	sheet = xfile.get_sheet_by_name('Sheet1')
	sheet['A2'] = 'Hello world afsdfkh asdkfh asd asdf asdf asf asdf'
	xfile.save('text2.xlsx')
	return HttpResponse(openpyxl.save_virtual_workbook(xfile), content_type='application/vnd.ms-excel')
	#response = HttpResponse(content_type='application/vnd.ms-excel')
	#response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
	#xlsx_data = WriteToExcel()
	#response.write(xlsx_data)
	#return response
'''

'''
def WriteToExcel():
	output = StringIO.StringIO()
	workbook = xlsxwriter.Workbook(output)
	worksheet = workbook.add_worksheet("Summery")
	worksheet.set_column('F:G', 12)

	title = workbook.add_format({
    	'bold': True,
    	'font_size': 14,
    	'align': 'center',
    	'valign': 'vcenter'
    	})

	header = workbook.add_format({
    	'bg_color': '#F7F7F7',
    	'color': 'black',
    	'align': 'center',
    	'valign': 'top',
    	'border': 1
    	})
	body = workbook.add_format({
		'border' : 1,
		'color': 'black',
    	'align': 'left'
		})
	body_body = workbook.add_format({
		'border' : 1,
		'color': 'black',
    	'align': 'left',
    	'bold' : True
		})

	title_text = u"2016 оны 01 сарын санхүүгийн зардалын мэдээ"
	worksheet.merge_range('C1:J1', title_text, title)
	worksheet.write('B3', u'№', header)
	for i in range(22):
		worksheet.write('B%s' %(i+4), i+1, body)
	worksheet.write('B27', 23, body)
	worksheet.write('B28', 24, body)
	worksheet.write('B31', 25, body)
	worksheet.write('B32', 26, body)
	worksheet.write('B33', 27, body)

	worksheet.merge_range('C3:G3', u'Зардалын төрлүүд', header)
	worksheet.merge_range('C4:G4', u'Үндсэн түүхий эд усны', body)
	worksheet.merge_range('C5:G5', u'Үндсэн ба нэмэгдэл цалин', body)
	worksheet.merge_range('C6:G6', u'Байгууллагаас төлөх ЭМД, НДШ', body)
	worksheet.merge_range('C7:G7', u'Цахилгаан - өөрийн хэрэглээ', body)
	worksheet.merge_range('C8:G8', u'Дулааг - өөрийн хэрэглээ', body)
	worksheet.merge_range('C9:G9', u'Ус - өөрийн хэрэглээ', body)
	worksheet.merge_range('C10:G10', u'Түлш шатахуун, тээврийн зардал', body)
	worksheet.merge_range('C11:G11', u'Харуул хамгаалалтын зардал', body)
	worksheet.merge_range('C12:G12', u'Засвар үйлчилгээний зардал', body)
	worksheet.merge_range('C13:G13', u'Ариутгал халдваргүйжүүлэлтийн зардал', body)
	worksheet.merge_range('C14:G14', u'Конторын хангамжийн зардал', body)
	worksheet.merge_range('C15:G15', u'Хөдөлмөр хамгааллын зардал', body)
	worksheet.merge_range('C16:G16', u'Маркетинг, борлуулалтын зардал', body)
	worksheet.merge_range('C17:G17', u'Лаборатори, шинжилшээний зардал', body)
	worksheet.merge_range('C18:G18', u'Гүйцэтгэх удирдлагын зардал', body)
	worksheet.merge_range('C19:G19', u'ТУЗ-н зардал', body)
	worksheet.merge_range('C20:G20', u'Үндсэн хөрөнгийн элэгдэл зардал', body)
	worksheet.merge_range('C21:G21', u'Зохицуулах үйлчилгээний хөлс', body)
	worksheet.merge_range('C22:G22', u'Гадны үйлчилгээний хөлс', body)
	worksheet.merge_range('C23:G23', u'Татварын зардал', body)
	worksheet.merge_range('C24:G24', u'Даатгалын зардал', body)
	worksheet.merge_range('C25:G25', u'Үндсэн үйл ажиллагааны бусад зардал', body)
	worksheet.merge_range('B26:G26', u'Үндсэн үйл ажиллагааны зардалын дүн', body_body)
	worksheet.merge_range('C27:G27', u'Ажиллагсдын нийгмийн зардалд', body)
	worksheet.merge_range('C28:G28', u'Хөрөнгө оруулалтын зардалд', body)
	worksheet.merge_range('B29:G29', u'Хөрөнгө оруулалтын өгөөжийн дүн', body_body)
	worksheet.merge_range('B30:G30', u'БОРЛУУЛАЛТЫН ОРЛОГЫН ШААРДАГДАХ ХЭМЖЭЭ / БОШ /', body_body)
	worksheet.merge_range('C31:G31', u'Шагнал урамшуулал', body)
	worksheet.merge_range('C32:G32', u'Спорт соёлын арга хэмжээ', body)
	worksheet.merge_range('C33:G33', u'Бусад', body)
	worksheet.merge_range('B34:G34', u'Үндсэн бус үйл ажиллагааны зардалын дүн', body_body)
	worksheet.merge_range('B35:G35', u'УХАТ-н үйл ажиллагааны нийт Зардлын дүн', body_body)

	worksheet.merge_range('H3:I3', u'Хэмжих нэгж', header)
	for i in range(32):
		worksheet.merge_range('H%s:I%s'%(i+4, i+4), u'мян.төгрөг', body)
	worksheet.merge_range('J3:K3', u'Үр дүн', header)

	for i in  range(32):
		worksheet.merge_range('J%s:K%s' %(i+4, i+4), u'', header)
	workbook.close()
	xlsx_data = output.getvalue()
	return xlsx_data
'''


def iter_all_strings():
	from string import ascii_uppercase
	import itertools
	size = 1
	while True:
		for s in itertools.product(ascii_uppercase, repeat=size):
			yield "".join(s)
		size +=1
def string():
	a = []
	for s in iter_all_strings():
		a.append(s)
		if s == 'ZZ':
			break
	return a

def excel_export(request, id = 0):
	tailan = SariinTailan.objects.get(id = id)
	xfile = openpyxl.load_workbook(os.path.join(BASE_DIR, 'table/sar.xlsx'))
	#sheet = xfile.get_sheet_by_name(u'4а, 4б -Орлогын задаргаа сараар')
	'''
	sheet1 = xfile.get_sheet_by_name(u'12- ТН+13-Дуудлага')
	date = u'%s оны %s сар' %(tailan.year, tailan.month)
	sheet1['D3'] = date #u'%s оны %s сар' %(tailan.year, tailan.month)
	sheet1['D5'] = tailan.tehnik_nohtsol.tehnik_nohtsol.all()[0].too
	sheet1['D6'] = tailan.tehnik_nohtsol.tehnik_nohtsol.all()[0].us
	sheet1['D7'] = tailan.tehnik_nohtsol.tehnik_nohtsol.all()[1].too
	sheet1['D8'] = float(tailan.tehnik_nohtsol.tehnik_nohtsol.all()[1].us)
	sheet1['D9'] = tailan.tehnik_nohtsol.sum0
	sheet1['D10'] = tailan.tehnik_nohtsol.sum1

	sheet1['C13'] = date
	sheet1['C15'] = tailan.sanal_gomdol.sum0
	sheet1['D15'] = tailan.sanal_gomdol.sum1
	sheet1['C16'] = tailan.sanal_gomdol.sanal_gomdol.all()[0].huleen_avsan
	sheet1['C17'] = tailan.sanal_gomdol.sanal_gomdol.all()[1].huleen_avsan
	sheet1['C18'] = tailan.sanal_gomdol.sanal_gomdol.all()[2].huleen_avsan
	sheet1['C19'] = tailan.sanal_gomdol.sanal_gomdol.all()[3].huleen_avsan
	sheet1['C20'] = tailan.sanal_gomdol.sanal_gomdol.all()[4].huleen_avsan
	sheet1['D16'] = tailan.sanal_gomdol.sanal_gomdol.all()[0].shiidverlesen
	sheet1['D17'] = tailan.sanal_gomdol.sanal_gomdol.all()[1].shiidverlesen
	sheet1['D18'] = tailan.sanal_gomdol.sanal_gomdol.all()[2].shiidverlesen
	sheet1['D19'] = tailan.sanal_gomdol.sanal_gomdol.all()[3].shiidverlesen
	sheet1['D20'] = tailan.sanal_gomdol.sanal_gomdol.all()[4].shiidverlesen
	#sheet6 = xfile.get_sheet_by_name(u'12- ТН+13-Дуудлага')
	sheet2 = xfile.get_sheet_by_name(u'11-Усны тасалдлын мэдээ')
	sheet2['A5'] = date
	sheet2['E5'] = tailan.tasaldal.tasaldal.all()[0].duration
	sheet2['F5'] = tailan.tasaldal.tasaldal.all()[1].duration
	sheet2['G5'] = tailan.tasaldal.tasaldal.all()[2].duration
	sheet2['H5'] = tailan.tasaldal.tasaldal.all()[3].duration
	sheet2['I5'] = tailan.tasaldal.tasaldal.all()[4].duration
	sheet2['J5'] = tailan.tasaldal.tasaldal.all()[5].duration
	sheet2['K5'] = tailan.tasaldal.tasaldal.all()[6].duration
	sheet2['L5'] = tailan.tasaldal.sum0

	sheet2['E6'] = tailan.tasaldal.tasaldal.all()[0].too
	sheet2['F6'] = tailan.tasaldal.tasaldal.all()[1].too
	sheet2['G6'] = tailan.tasaldal.tasaldal.all()[2].too
	sheet2['H6'] = tailan.tasaldal.tasaldal.all()[3].too
	sheet2['I6'] = tailan.tasaldal.tasaldal.all()[4].too
	sheet2['J6'] = tailan.tasaldal.tasaldal.all()[5].too
	sheet2['K6'] = tailan.tasaldal.tasaldal.all()[6].too
	sheet2['K6'] = tailan.tasaldal.sum1

	sheet3 = xfile.get_sheet_by_name(u'9 - ААН-голчийн судалгаа')
	s3 = 5
	for too, i in enumerate(tailan.golch.all()[0].buteegdehuun.all()):
		sheet3['C%s' %s3] = i.too
		sheet3['E%s' %s3] = tailan.example(too)
		sheet3['F%s' %s3] = tailan.example_12(too)
		s3 += 1
	'''
	a = 4
	b = 5
	c = 4
	d = 3
	sheet4 = xfile.get_sheet_by_name(u'8 - ХЭРЭГЛЭГЧ-СУУРЬ ҮНЭ')
	sheet5 = xfile.get_sheet_by_name(u'4а, 4б -Орлогын задаргаа сараар')
	sheet6 = xfile.get_sheet_by_name(u'5а - зардал сараар')

	sheet5_7 = sheet5_8 = sheet5_9 = sheet5_10 = sheet5_11 = sheet5_12 = sheet5_13 = sheet5_14 = sheet5_15 = sheet5_17 = sheet5_18 = sheet5_19 = sheet5_20 = sheet5_21 = sheet5_22 = sheet5_23 = sheet5_24 = sheet5_37 = sheet5_38 = sheet5_39 = sheet5_40 = sheet5_41 = sheet5_42 = sheet5_43 = sheet5_44 = sheet5_47 = sheet5_48 = sheet5_49 = sheet5_50 = sheet5_51 = sheet5_52 = sheet5_53 = sheet5_55 = sheet5_56= 0
	sheet6_52 = sheet6_51 = sheet6_50 = sheet6_46 = sheet6_47 = sheet6_42 = sheet6_43 = 0
	sheet6_6 = sheet6_7 = sheet6_8 = sheet6_9 = sheet6_10 = sheet6_11 = sheet6_12 = 0
	sheet6_13 = sheet6_14 = sheet6_15 = sheet6_16 = sheet6_17 = sheet6_18 = sheet6_19 = 0
	sheet6_20 = sheet6_21 = sheet6_22 = sheet6_23 = sheet6_24 = sheet6_25 = sheet6_26 = 0
	sheet6_28 = sheet6_29 = sheet6_32 = sheet6_33 = sheet6_34 = 0
	for i in range(tailan.month):
		
		st = SariinTailan.objects.get(tze = tailan.tze, year = tailan.year, month = i+1)
		
		sheet4['%s6' %string()[a]] = st.hereglegch.niit_tsever_count()
		sheet4['%s6' %string()[b]] = st.hereglegch.niit_bohir_count()

		sheet4['%s7' %string()[a]] = st.hereglegch.hereglegch.all()[0].htoo
		sheet4['%s7' %string()[b]] = st.hereglegch.hereglegch.all()[1].htoo

		sheet4['%s8' %string()[a]] = st.hereglegch.hereglegch.all()[2].htoo
		sheet4['%s8' %string()[b]] = st.hereglegch.hereglegch.all()[3].htoo
		
		sheet4['%s9' %string()[a]] = st.hereglegch.hereglegch.all()[4].htoo
		sheet4['%s9' %string()[b]] = st.hereglegch.hereglegch.all()[5].htoo
		
		sheet4['%s10' %string()[a]] = st.hereglegch.hereglegch.all()[6].htoo
		sheet4['%s10' %string()[b]] = st.hereglegch.hereglegch.all()[7].htoo
		
		sheet4['%s11' %string()[a]] = st.hereglegch.hereglegch.all()[8].htoo
		sheet4['%s11' %string()[b]] = st.hereglegch.hereglegch.all()[9].htoo


		sheet4['%s12' %string()[a]] = st.hereglegch.aanb_tsever_count()
		sheet4['%s12' %string()[b]] = st.hereglegch.aanb_bohir_count()

		sheet4['%s13' %string()[a]] = st.hereglegch.hereglegch.all()[10].htoo
		sheet4['%s13' %string()[b]] = st.hereglegch.hereglegch.all()[11].htoo

		sheet4['%s14' %string()[a]] = st.hereglegch.hereglegch.all()[12].htoo
		sheet4['%s14' %string()[b]] = st.hereglegch.hereglegch.all()[13].htoo

		sheet4['%s15' %string()[a]] = st.hereglegch.utb_zoovor_tsever_count()
		sheet4['%s15' %string()[b]] = st.hereglegch.utb_zoovor_bohir_count()

		sheet4['%s16' %string()[a]] = st.hereglegch.ahuin_tsever_count()
		sheet4['%s16' %string()[b]] = st.hereglegch.ahuin_bohir_count()

		sheet4['%s17' %string()[a]] = st.orlogo.all_tsever_count()
		sheet4['%s17' %string()[b]] = st.orlogo.all_bohir_count()

		sheet4['%s18' %string()[a]] = st.orlogo.buteegdehuun.all()[2].too
		sheet4['%s18' %string()[b]] = st.orlogo.buteegdehuun.all()[3].too

		sheet4['%s19' %string()[a]] = st.orlogo.buteegdehuun.all()[4].too
		sheet4['%s19' %string()[b]] = st.orlogo.buteegdehuun.all()[5].too

		sheet4['%s20' %string()[a]] = st.orlogo.buteegdehuun.all()[6].too
		sheet4['%s20' %string()[b]] = st.orlogo.buteegdehuun.all()[7].too

		sheet4['%s21' %string()[a]] = st.orlogo.buteegdehuun.all()[8].too
		sheet4['%s21' %string()[b]] = st.orlogo.buteegdehuun.all()[9].too

		sheet4['%s22' %string()[a]] = st.orlogo.aanb_tsever_count()
		sheet4['%s22' %string()[b]] = st.orlogo.aanb_bohir_count()

		sheet4['%s23' %string()[a]] = st.orlogo.buteegdehuun.all()[10].too
		sheet4['%s23' %string()[b]] = st.orlogo.buteegdehuun.all()[11].too

		sheet4['%s24' %string()[a]] = st.orlogo.buteegdehuun.all()[12].too
		sheet4['%s24' %string()[b]] = st.orlogo.buteegdehuun.all()[13].too

		sheet4['%s25' %string()[a]] = st.orlogo.utbz_tsever_count()
		sheet4['%s25' %string()[b]] = st.orlogo.utbz_bohir_count()

		sheet4['%s26' %string()[a]] = st.orlogo.ahuin_tsever_count()
		sheet4['%s26' %string()[b]] = st.orlogo.ahuin_bohir_count()

		#for s5too,d in enumerate(st.orlogo.buteegdehuun.all()):
		sheet5['%s7' %string()[c]] = st.orlogo.buteegdehuun.all()[0].too
		sheet5['%s8' %string()[c]] = st.orlogo.buteegdehuun.all()[2].too
		sheet5['%s9' %string()[c]] = st.orlogo.buteegdehuun.all()[4].too
		sheet5['%s10' %string()[c]] = st.orlogo.buteegdehuun.all()[6].too
		sheet5['%s11' %string()[c]] = st.orlogo.buteegdehuun.all()[8].too
		sheet5['%s12' %string()[c]] = st.orlogo.buteegdehuun.all()[10].too
		sheet5['%s13' %string()[c]] = st.orlogo.buteegdehuun.all()[12].too
		sheet5['%s14' %string()[c]] = st.orlogo.buteegdehuun.all()[14].too
		sheet5['%s15' %string()[c]] = st.orlogo.buteegdehuun.all()[16].too


		sheet5['%s17' %string()[c]] = st.orlogo.buteegdehuun.all()[1].too
		sheet5['%s18' %string()[c]] = st.orlogo.buteegdehuun.all()[3].too
		sheet5['%s19' %string()[c]] = st.orlogo.buteegdehuun.all()[5].too
		sheet5['%s20' %string()[c]] = st.orlogo.buteegdehuun.all()[7].too
		sheet5['%s21' %string()[c]] = st.orlogo.buteegdehuun.all()[9].too
		sheet5['%s22' %string()[c]] = st.orlogo.buteegdehuun.all()[11].too
		sheet5['%s23' %string()[c]] = st.orlogo.buteegdehuun.all()[13].too
		sheet5['%s24' %string()[c]] = st.orlogo.buteegdehuun.all()[17].too

		#sheet5['%s35' %string()[c]] = st.orlogo_une()[0]
		sheet5['%s37' %string()[c]] = st.orlogo_une()[0]
		sheet5['%s38' %string()[c]] = st.orlogo_une()[2]
		sheet5['%s39' %string()[c]] = st.orlogo_une()[4]
		sheet5['%s40' %string()[c]] = st.orlogo_une()[6]
		sheet5['%s41' %string()[c]] = st.orlogo_une()[8]
		sheet5['%s42' %string()[c]] = st.orlogo_une()[10]
		sheet5['%s43' %string()[c]] = st.orlogo_une()[12]
		sheet5['%s44' %string()[c]] = st.orlogo_une()[14]

		#sheet5['%s45' %string()[c]] = st.orlogo_une()[1]
		sheet5['%s47' %string()[c]] = st.orlogo_une()[1]
		sheet5['%s48' %string()[c]] = st.orlogo_une()[3]
		sheet5['%s49' %string()[c]] = st.orlogo_une()[5]
		sheet5['%s50' %string()[c]] = st.orlogo_une()[7]
		sheet5['%s51' %string()[c]] = st.orlogo_une()[9]
		sheet5['%s52' %string()[c]] = st.orlogo_une()[11]
		sheet5['%s53' %string()[c]] = st.orlogo_une()[15]
		sheet5['%s55' %string()[c]] = st.suuri_une_count()
		sheet5['%s56' %string()[c]] = st.tariff_hereglegch_us().suuri_une

		sheet5_7 += sheet5['%s7' %string()[c]].value
		sheet5_8 += sheet5['%s7' %string()[c]].value
		sheet5_9 += sheet5['%s7' %string()[c]].value
		sheet5_10 += sheet5['%s7' %string()[c]].value
		sheet5_11 += sheet5['%s7' %string()[c]].value
		sheet5_12 += sheet5['%s7' %string()[c]].value
		sheet5_13 += sheet5['%s7' %string()[c]].value
		sheet5_14+= sheet5['%s7' %string()[c]].value
		sheet5_15 += sheet5['%s7' %string()[c]].value
		sheet5_17 += sheet5['%s7' %string()[c]].value
		sheet5_18 += sheet5['%s7' %string()[c]].value
		sheet5_19 += sheet5['%s7' %string()[c]].value
		sheet5_20 += sheet5['%s7' %string()[c]].value
		sheet5_21 += sheet5['%s7' %string()[c]].value
		sheet5_22 += sheet5['%s7' %string()[c]].value
		sheet5_23 += sheet5['%s7' %string()[c]].value
		sheet5_24 += sheet5['%s7' %string()[c]].value
		sheet5_37 += sheet5['%s7' %string()[c]].value
		sheet5_38 += sheet5['%s7' %string()[c]].value
		sheet5_39 += sheet5['%s7' %string()[c]].value
		sheet5_40 += sheet5['%s7' %string()[c]].value
		sheet5_41 += sheet5['%s7' %string()[c]].value
		sheet5_42 += sheet5['%s7' %string()[c]].value
		sheet5_43 += sheet5['%s7' %string()[c]].value
		sheet5_44 += sheet5['%s7' %string()[c]].value
		sheet5_47 += sheet5['%s7' %string()[c]].value
		sheet5_48 += sheet5['%s7' %string()[c]].value
		sheet5_49 += sheet5['%s7' %string()[c]].value
		sheet5_50 += sheet5['%s7' %string()[c]].value
		sheet5_51 += sheet5['%s7' %string()[c]].value
		sheet5_52 += sheet5['%s7' %string()[c]].value
		sheet5_53 += sheet5['%s7' %string()[c]].value
		sheet5_55 += sheet5['%s7' %string()[c]].value
		sheet5_56 += sheet5['%s7' %string()[c]].value
		#sheet4['C29'] = tailan.orlogo.all_orh_count()

		#sheet4['D29'] = tailan.tariff_hereglegch_us().suuri_une

		#sheet4['E29'] = tailan.suuri_une_count()

		#sheet4['F29'] = tailan.suuri_12_count()

		sheet6['%s6' %string()[d]] = st.zardal.z2.nem0
		if sheet6['%s6' %string()[d]].value:
			sheet6_6 += sheet6['%s6' %string()[d]].value

		sheet6['%s7' %string()[d]] = st.zardal.z2.nem1
		if sheet6['%s7' %string()[d]].value:
			sheet6_7 += sheet6['%s7' %string()[d]].value

		sheet6['%s8' %string()[d]] = st.zardal.z3.nem0
		sheet6_8 += sheet6['%s8' %string()[d]].value

		sheet6['%s9' %string()[d]] = st.zardal.z3.nem1
		sheet6_9 += sheet6['%s9' %string()[d]].value

		sheet6['%s10' %string()[d]] = st.zardal.z3.nem2
		sheet6_10 += sheet6['%s10' %string()[d]].value

		sheet6['%s11' %string()[d]] = st.zardal.z3.nem4
		sheet6_11 += sheet6['%s11' %string()[d]].value

		sheet6['%s12' %string()[d]] = st.zardal.z3.nem6
		sheet6_12 += sheet6['%s12' %string()[d]].value

		sheet6['%s13' %string()[d]] = st.zardal.z4.nem
		sheet6_13 += sheet6['%s13' %string()[d]].value

		sheet6['%s14' %string()[d]] = st.zardal.z5.nem
		sheet6_14 += sheet6['%s14' %string()[d]].value

		sheet6['%s15' %string()[d]] = st.zardal.z6.nem
		sheet6_15 += sheet6['%s15' %string()[d]].value

		sheet6['%s16' %string()[d]] = st.zardal.z7.nem
		sheet6_16 += sheet6['%s16' %string()[d]].value

		sheet6['%s17' %string()[d]] = st.zardal.z8.nem
		sheet6_17 += sheet6['%s17' %string()[d]].value

		sheet6['%s18' %string()[d]] = st.zardal.z9.nem
		sheet6_18 += sheet6['%s18' %string()[d]].value

		sheet6['%s19' %string()[d]] = st.zardal.z10.nem
		sheet6_19 += sheet6['%s19' %string()[d]].value

		sheet6['%s20' %string()[d]] = st.zardal.z11.nem
		sheet6_20 += sheet6['%s20' %string()[d]].value

		sheet6['%s21' %string()[d]] = st.zardal.z12.nem
		sheet6_21 += sheet6['%s21' %string()[d]].value

		sheet6['%s22' %string()[d]] = st.zardal.z13.nem0
		sheet6_22 += sheet6['%s22' %string()[d]].value

		sheet6['%s23' %string()[d]] = st.zardal.z13.nem
		sheet6_23 += sheet6['%s23' %string()[d]].value

		sheet6['%s24' %string()[d]] = st.zardal.z14.nem
		sheet6_24 += sheet6['%s24' %string()[d]].value

		sheet6['%s25' %string()[d]] = st.zardal.z15.nem
		sheet6_25 += sheet6['%s25' %string()[d]].value

		if st.zardal.z16.count:
			sheet6['%s26' %string()[d]] = st.zardal.z16.count
		else:
			sheet6['%s26' %string()[d]] = 0
		sheet6_26 += sheet6['%s26' %string()[d]].value
		
		sheet6['%s28' %string()[d]] = st.zardal.z17.nem
		sheet6_28 += sheet6['%s28' %string()[d]].value

		sheet6['%s29' %string()[d]] = st.zardal.z18.nem
		sheet6_29 += sheet6['%s29' %string()[d]].value

		sheet6['%s32' %string()[d]] = st.zardal.z19.total0
		sheet6_32 += sheet6['%s32' %string()[d]].value

		sheet6['%s33' %string()[d]] = st.zardal.z19.total1
		sheet6_33 += sheet6['%s33' %string()[d]].value

		sheet6['%s34' %string()[d]] = st.zardal.z19.total2
		sheet6_34 += sheet6['%s34' %string()[d]].value

		sheet6['%s42' %string()[d]] = st.zardal.z1.undsen_material.all()[0].undsen_tuuhii_ed_usnii
		sheet6_42 += sheet6['%s42' %string()[d]].value

		sheet6['%s43' %string()[d]] = st.zardal.z1.undsen_material.all()[1].undsen_tuuhii_ed_usnii
		sheet6_43 += sheet6['%s43' %string()[d]].value


		sheet6['%s46' %string()[d]] = st.olborlolt.tsever
		sheet6_46 += sheet6['%s46' %string()[d]].value

		sheet6['%s47' %string()[d]] = st.olborlolt.bohir
		sheet6_47 += sheet6['%s47' %string()[d]].value

		sheet6['%s50' %string()[d]] = sheet6['%s42' %string()[d]].value * sheet6['%s46' %string()[d]].value
		sheet6_50 += sheet6['%s50' %string()[d]].value

		sheet6['%s51' %string()[d]] = sheet6['%s43' %string()[d]].value * sheet6['%s47' %string()[d]].value
		sheet6_51 += sheet6['%s51' %string()[d]].value

		sheet6['%s52' %string()[d]] = st.zardal1_suuri_une()
		sheet6_52 += sheet6['%s52' %string()[d]].value





		a += 2
		b += 2
		c += 1
		d += 1

	sheet5['%s7' %string()[16]] = sheet5_7
	sheet5['%s8' %string()[16]] = sheet5_8 
	sheet5['%s9' %string()[16]] = sheet5_9
	sheet5['%s10' %string()[16]] = sheet5_10
	sheet5['%s11' %string()[16]] = sheet5_11
	sheet5['%s12' %string()[16]] = sheet5_12
	sheet5['%s13' %string()[16]] = sheet5_13
	sheet5['%s14' %string()[16]] = sheet5_14
	sheet5['%s15' %string()[16]] = sheet5_15


	sheet5['%s17' %string()[16]] = sheet5_17
	sheet5['%s18' %string()[16]] = sheet5_18
	sheet5['%s19' %string()[16]] = sheet5_19
	sheet5['%s20' %string()[16]] = sheet5_20
	sheet5['%s21' %string()[16]] = sheet5_21
	sheet5['%s22' %string()[16]] = sheet5_22
	sheet5['%s23' %string()[16]] = sheet5_23
	sheet5['%s24' %string()[16]] = sheet5_24

	#sheet5['%s35' %string()[16]] = st.orlogo_une()[0]
	sheet5['%s37' %string()[16]] = sheet5_37
	sheet5['%s38' %string()[16]] = sheet5_38
	sheet5['%s39' %string()[16]] = sheet5_39
	sheet5['%s40' %string()[16]] = sheet5_40
	sheet5['%s41' %string()[16]] = sheet5_41
	sheet5['%s42' %string()[16]] = sheet5_42
	sheet5['%s43' %string()[16]] = sheet5_43
	sheet5['%s44' %string()[16]] = sheet5_44

	#sheet5['%s45' %string()[16]] = 
	sheet5['%s47' %string()[16]] = sheet5_47
	sheet5['%s48' %string()[16]] = sheet5_48
	sheet5['%s49' %string()[16]] = sheet5_49
	sheet5['%s50' %string()[16]] = sheet5_50
	sheet5['%s51' %string()[16]] = sheet5_51
	sheet5['%s52' %string()[16]] = sheet5_52
	sheet5['%s53' %string()[16]] = sheet5_53
	sheet5['%s55' %string()[16]] = sheet5_55
	sheet5['%s56' %string()[16]] = sheet5_56

	sheet6['%s6' %string()[15]] = sheet6_6
	sheet6['%s7' %string()[15]] = sheet6_7
	sheet6['%s8' %string()[15]] = sheet6_8
	sheet6['%s9' %string()[15]] = sheet6_9
	sheet6['%s10' %string()[15]] = sheet6_10
	sheet6['%s11' %string()[15]] = sheet6_11
	sheet6['%s12' %string()[15]] = sheet6_12
	sheet6['%s13' %string()[15]] = sheet6_13
	sheet6['%s14' %string()[15]] = sheet6_14
	sheet6['%s15' %string()[15]] = sheet6_15
	sheet6['%s16' %string()[15]] = sheet6_16
	sheet6['%s17' %string()[15]] = sheet6_17
	sheet6['%s18' %string()[15]] = sheet6_18
	sheet6['%s19' %string()[15]] = sheet6_19
	sheet6['%s20' %string()[15]] = sheet6_20
	sheet6['%s21' %string()[15]] = sheet6_21
	sheet6['%s22' %string()[15]] = sheet6_22
	sheet6['%s23' %string()[15]] = sheet6_23
	sheet6['%s24' %string()[15]] = sheet6_24
	sheet6['%s25' %string()[15]] = sheet6_25
	sheet6['%s26' %string()[15]] = sheet6_26
	sheet6['%s28' %string()[15]] = sheet6_28
	sheet6['%s29' %string()[15]] = sheet6_29
	sheet6['%s32' %string()[15]] = sheet6_32
	sheet6['%s33' %string()[15]] = sheet6_33
	sheet6['%s34' %string()[15]] = sheet6_34


	sheet6['%s42' %string()[15]] = sheet6_42
	sheet6['%s43' %string()[15]] = sheet6_43
	sheet6['%s46' %string()[15]] = sheet6_46
	sheet6['%s47' %string()[15]] = sheet6_47
	sheet6['%s50' %string()[15]] = sheet6_50
	sheet6['%s51' %string()[15]] = sheet6_51
	sheet6['%s52' %string()[15]] = sheet6_52


		

	'''
	sheet5 = xfile.get_sheet_by_name(u'7 - ӨРТӨГ')

	for sh5, i in enumerate("DEFG"):
		sheet5['%s11' %i] = tailan.zardal.z2.tsalin.all()
		print i
		print sh5

	'''
	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=report.xlsx'
	xfile.save(response)
	return response




















class HereglegchFormviewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/hereglegchlist.html'

class AANBFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/golchiinsuuriune.html'

class UsHudaldanAvahFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/ushudaldanavahlist.html'

class GolchFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/golchlist.html'

class TariffFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/tarifflist.html'

	def get_context_data(self, *args, **kwargs):
		context = super(TariffFormViewList, self).get_context_data(*args, **kwargs)
		try: context['tariffus'] = get_object_or_none(self.tailan.tariff_hereglegch.get(tze = self.baiguullaga, name =0))
		except: pass
		return context

class GolchHangagchFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/golchhangagchlist.html'

class OrlogoFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/orlogolist.html'

class BusadOrlogoFormViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/busadorlogolist.html'

class ZardalFormView1List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist1.html'

class ZardalFormView2List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist2.html'

class ZardalFormView3List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist3.html'

class ZardalFormView4List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist4.html'

class ZardalFormView5List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist5.html'

class ZardalFormView6List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist6.html'

class ZardalFormView7List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist7.html'

class ZardalFormView8List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist8.html'

class ZardalFormView9List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist9.html'

class ZardalFormView10List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist10.html'

class ZardalFormView11List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist11.html'

class ZardalFormView12List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist12.html'

class ZardalFormView13List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist13.html'

class ZardalFormView14List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist14.html'

class ZardalFormView15List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist15.html'

class ZardalFormView16List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist16.html'

class ZardalFormView17List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist17.html'

class ZardalFormView18List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist18.html'

class ZardalFormView19List(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/zardallist19.html'

class SanalGomdolViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/sanalgomdol.html'

class TehnikNohtsolViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/tehniknohtsol.html'

class TasaldalViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/tasaldal.html'

class TsalinViewList(LoginRequired, TailanView, TemplateView):
	template_name = 'listuud/tsalinlist.html'
